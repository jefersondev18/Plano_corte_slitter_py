"""
Plano de Corte — Otimizador de Combinações de Matrizes
=======================================================
Regras de negócio:
  1. O usuário escolhe: espessura, tipo de material e uma MATRIZ ÂNCORA
  2. A MATRIZ ÂNCORA é OBRIGATÓRIA no plano (N >= 1 corte)
  3. A LARGURA DA BOBINA é fixa: 1200 mm (padrão), 1000 mm ou 1500 mm
     → O script tenta 1200 primeiro; se não houver combinação válida, tenta 1000, depois 1500
  4. Complementares: outras matrizes da mesma espessura + tipo que preenchem o espaço restante
  5. Uma combinação é VÁLIDA quando:
        perda_mm  = largura_bobina - Σ (Desenvolvimento_i × N_cortes_i)
        perda_%   = perda_mm / largura_bobina × 100
        0,67% ≤ perda_% ≤ 1,70%
  6. LIMITE DE CORTES (opcional): o usuário pode informar um número máximo para a SOMA total
     de cortes de uma combinação (restrição de máquina). Se vazio, sem limite.
  7. QUANTIDADE DE KG por combinação (calculada por matriz):
        Peso_médio_bobina = Peso_informado / Qtd_bobinas  (se não informado: 12.000 / 1 = 12.000 kg)
        KG_i = (Peso_médio_bobina / Largura_bobina) × (N_cortes_i × Desenvolvimento_i × Qtd_bobinas)
"""

import os
import platform
import pandas as pd
from itertools import combinations, product as iproduct

# ──────────────────────────────────────────────
#  PATHS
# ──────────────────────────────────────────────
SO = platform.system()
if SO == 'Windows':
    BASE_INPUT  = r'C:\Users\marce\OneDrive\Documentos\GitHub\plano_corte\input'
    BASE_OUTPUT = r'C:\Users\marce\OneDrive\Documentos\GitHub\plano_corte\output'
elif SO == 'Linux':
    BASE_INPUT  = r'/home/stark/Documentos/Dev/Plano_corte_py/files/input'
    BASE_OUTPUT = r'/home/stark/Documentos/Dev/Plano_corte_py/files/output'
else:
    raise Exception(f'Sistema Operacional não suportado: {SO}')

# ──────────────────────────────────────────────
#  PARÂMETROS DE NEGÓCIO
# ──────────────────────────────────────────────
LARGURAS_BOBINA     = [1200, 1000, 1500]   # ordem de tentativa
PERDA_MIN_PCT       = 0.67                 # % mínimo de perda aceito
PERDA_MAX_PCT       = 1.70                 # % máximo de perda aceito
MAX_COMP_NA_COMBO   = 2                    # máx de matrizes COMPLEMENTARES por combinação

# REFILO — regras por espessura
REFILO_MIN_ATE_3MM  = 10                   # mm mínimo de refilo para espessuras ≤ 3.0 mm
REFILO_MIN_ACIMA_3MM = 14                  # mm mínimo de refilo para espessuras > 3.0 mm

# KG — padrões para o cálculo de quantidade
PESO_MEDIO_BOB_PAD  = 12_000               # kg (12 ton) — pode ser sobrescrito pelo usuário
QTD_BOBINAS_PAD     = 1                    # quantidade padrão se usuário não informar


# ──────────────────────────────────────────────
#  CARGA E LIMPEZA
# ──────────────────────────────────────────────
def carregar_dados(caminho: str) -> pd.DataFrame:
    df = pd.read_excel(caminho)
    df['Matriz']           = df['Matriz'].astype(str).str.strip()
    df['Tipo de material'] = df['Tipo de material'].astype(str).str.strip()
    df['Espessura']        = pd.to_numeric(df['Espessura'], errors='coerce')
    df['Desenvolvimento']  = pd.to_numeric(df['Desenvolvimento'], errors='coerce')
    df = df.dropna(subset=['Espessura', 'Desenvolvimento', 'Matriz', 'Tipo de material'])
    df = df[df['Desenvolvimento'] > 0]
    df = df[~df['Matriz'].isin(['nan', ''])]
    return df


# ──────────────────────────────────────────────
#  HELPERS DE CONSULTA
# ──────────────────────────────────────────────
def listar_espessuras(df: pd.DataFrame) -> list[float]:
    return sorted(df['Espessura'].unique())


def listar_tipos(df: pd.DataFrame, espessura: float) -> list[str]:
    return sorted(df[df['Espessura'] == espessura]['Tipo de material'].unique())


def listar_matrizes(df: pd.DataFrame, espessura: float, tipo: str) -> pd.DataFrame:
    mask = (df['Espessura'] == espessura) & (df['Tipo de material'] == tipo)
    return (
        df[mask]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'Dev_mm'})
        .sort_values('Dev_mm', ascending=False)
        .reset_index(drop=True)
    )


def get_dev(df: pd.DataFrame, matriz: str, espessura: float) -> float:
    mask = (df['Matriz'] == matriz) & (df['Espessura'] == espessura)
    vals = df[mask]['Desenvolvimento'].dropna()
    if vals.empty:
        raise ValueError(f"Matriz '{matriz}' / esp {espessura} mm não encontrada.")
    return float(vals.mean())


# ──────────────────────────────────────────────
#  BUSCA PARA UMA LARGURA ESPECÍFICA
# ──────────────────────────────────────────────
def _buscar_para_largura(
    dev_ancora: float,
    ancora: str,
    mats_comp: list[str],
    devs_comp: list[float],
    largura: int,
    max_comp: int,
    limite_cortes: int | None = None,
    espessura: float = 0.0,
) -> list[dict]:
    """
    Busca todas as combinações válidas para uma largura de bobina específica.
    A âncora é sempre obrigatória (N >= 1). Complementares são opcionais.
    limite_cortes: soma máxima de todos os N_cortes da combinação (None = sem limite).
    espessura: usado para determinar o refilo mínimo obrigatório.
    Retorna lista de dicts com os resultados.
    
    Validação em cascata:
      1. Perda % deve estar entre PERDA_MIN_PCT e PERDA_MAX_PCT
      2. Perda mm deve ser >= refilo mínimo (10mm se esp≤3mm, 14mm se esp>3mm)
      3. Se passar ambas: Status = "✓ Válida"
      4. Se passar % mas não refilo: Status = "Fora da regra"
    """
    perda_min_mm = largura * PERDA_MIN_PCT / 100
    perda_max_mm = largura * PERDA_MAX_PCT / 100
    max_n_ancora = int(largura / dev_ancora)
    
    # Determina refilo mínimo pela espessura
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM

    resultados = []

    for n_ancora in range(1, max_n_ancora + 1):
        soma_ancora      = dev_ancora * n_ancora
        espaco_restante  = largura - soma_ancora
        if espaco_restante < 0:
            break

        # ── Caso: só a âncora ──
        perda_mm = espaco_restante
        total_cortes_ancora = n_ancora
        
        # Validação cascata: primeiro %, depois refilo
        passa_pct = (perda_min_mm <= perda_mm <= perda_max_mm)
        passa_refilo = (perda_mm >= refilo_min)
        passa_cortes = (limite_cortes is None or total_cortes_ancora <= limite_cortes)
        
        if passa_pct and passa_cortes:
            if passa_refilo:
                status = "✓ Válida"
            else:
                status = "Fora da regra"
            
            resultados.append({
                'Combinacao':      f'{ancora}(x{n_ancora})',
                'N_ancora':        n_ancora,
                'Num_comp':        0,
                'Total_cortes':    n_ancora,
                'Detalhes': [
                    {'Matriz': ancora, 'Desenvolvimento_mm': dev_ancora,
                     'N_cortes': n_ancora, 'Subtotal_mm': round(soma_ancora, 3)}
                ],
                'Soma_cortes_mm':  round(soma_ancora, 3),
                'Perda_mm':        round(perda_mm, 3),
                'Perda_pct':       round(perda_mm / largura * 100, 4),
                'Largura_bobina':  largura,
                'Status':          status,
            })

        # ── Caso: âncora + complementares ──
        if espaco_restante < min(devs_comp, default=largura + 1):
            continue  # nenhum complementar cabe

        comp_ok_idx = [i for i, d in enumerate(devs_comp) if d <= espaco_restante]
        if not comp_ok_idx:
            continue

        for tam in range(1, min(max_comp, len(comp_ok_idx)) + 1):
            for idxs in combinations(comp_ok_idx, tam):
                ds = [devs_comp[i] for i in idxs]
                ms = [mats_comp[i] for i in idxs]
                lims = [max(1, int(espaco_restante / d)) for d in ds]

                for ns in iproduct(*[range(1, lim + 1) for lim in lims]):
                    soma_comp    = sum(d * n for d, n in zip(ds, ns))
                    soma_total   = soma_ancora + soma_comp
                    total_cortes = n_ancora + sum(ns)

                    if soma_total > largura:
                        continue
                    if limite_cortes is not None and total_cortes > limite_cortes:
                        continue

                    perda_mm = largura - soma_total
                    
                    # Validação cascata
                    passa_pct = (perda_min_mm <= perda_mm <= perda_max_mm)
                    passa_refilo = (perda_mm >= refilo_min)
                    
                    if passa_pct:
                        if passa_refilo:
                            status = "✓ Válida"
                        else:
                            status = "Fora da regra"
                        
                        detalhes = [
                            {'Matriz': ancora, 'Desenvolvimento_mm': dev_ancora,
                             'N_cortes': n_ancora, 'Subtotal_mm': round(soma_ancora, 3)}
                        ] + [
                            {'Matriz': m, 'Desenvolvimento_mm': d,
                             'N_cortes': n, 'Subtotal_mm': round(d * n, 3)}
                            for m, d, n in zip(ms, ds, ns)
                        ]
                        comp_str = ' + '.join(f'{m}(x{n})' for m, n in zip(ms, ns))
                        resultados.append({
                            'Combinacao':      f'{ancora}(x{n_ancora}) + {comp_str}',
                            'N_ancora':        n_ancora,
                            'Num_comp':        tam,
                            'Total_cortes':    total_cortes,
                            'Detalhes':        detalhes,
                            'Soma_cortes_mm':  round(soma_total, 3),
                            'Perda_mm':        round(perda_mm, 3),
                            'Perda_pct':       round(perda_mm / largura * 100, 4),
                            'Largura_bobina':  largura,
                            'Status':          status,
                        })

    return resultados


# ──────────────────────────────────────────────
#  FUNÇÃO PRINCIPAL — TENTA LARGURAS EM ORDEM
# ──────────────────────────────────────────────
def encontrar_combinacoes(
    df: pd.DataFrame,
    espessura: float,
    tipo_material: str,
    matriz_ancora: str,
    max_comp: int = MAX_COMP_NA_COMBO,
    limite_cortes: int | None = None,
) -> tuple[pd.DataFrame, int]:
    """
    Tenta as larguras na ordem [1200, 1000, 1500].
    Para na primeira que retornar combinações válidas.
    limite_cortes: soma máxima de todos os N_cortes por combinação (None = sem limite).
    Retorna (DataFrame de resultados, largura_usada).
    """
    dev_ancora = get_dev(df, matriz_ancora, espessura)

    # Complementares: mesma espessura + tipo, excluindo âncora
    mask_comp = (
        (df['Espessura'] == espessura) &
        (df['Tipo de material'] == tipo_material) &
        (df['Matriz'] != matriz_ancora)
    )
    cands = (
        df[mask_comp]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'dev'})
        .sort_values('dev', ascending=False)
    )
    mats_comp = cands['Matriz'].tolist()
    devs_comp = cands['dev'].tolist()

    for largura in LARGURAS_BOBINA:
        print(f"  → Tentando largura {largura} mm ...", end=' ')

        # Dev âncora deve caber ao menos 1x na bobina
        if dev_ancora > largura:
            print(f"âncora ({dev_ancora:.1f}mm) não cabe. Pulando.")
            continue

        resultados = _buscar_para_largura(
            dev_ancora, matriz_ancora, mats_comp, devs_comp, largura, max_comp,
            limite_cortes=limite_cortes,
            espessura=espessura,
        )

        if resultados:
            print(f"{len(resultados)} combinações encontradas. ✓")
            df_res = (
                pd.DataFrame(resultados)
                .sort_values(['Perda_pct', 'N_ancora', 'Num_comp'])
                .reset_index(drop=True)
            )
            return df_res, largura
        else:
            print("nenhuma combinação válida.")

    return pd.DataFrame(), 0


# ──────────────────────────────────────────────
#  EXIBIÇÃO NO TERMINAL
# ──────────────────────────────────────────────
def exibir(df_res: pd.DataFrame, largura: int,
           ancora: str, esp: float, tipo: str,
           limite_cortes: int | None = None) -> None:
    sep = "=" * 76
    print(f"\n{sep}")
    print("  PLANO DE CORTE — COMBINAÇÕES VÁLIDAS")
    print(sep)
    print(f"  Âncora         : {ancora}")
    print(f"  Espessura      : {esp} mm")
    print(f"  Tipo material  : {tipo}")
    print(f"  Largura bobina : {largura} mm  (padrão usado)")
    print(f"  Janela de perda: {PERDA_MIN_PCT}% – {PERDA_MAX_PCT}%"
          f"  |  {largura * PERDA_MIN_PCT / 100:.2f} mm – {largura * PERDA_MAX_PCT / 100:.2f} mm")
    
    refilo_min = REFILO_MIN_ATE_3MM if esp <= 3.0 else REFILO_MIN_ACIMA_3MM
    print(f"  Refilo mínimo  : {refilo_min} mm  (regra para esp {'≤' if esp <= 3.0 else '>'} 3.0 mm)")
    
    if limite_cortes is not None:
        print(f"  Limite cortes  : {limite_cortes} cortes (soma total por combinação)")

    if df_res.empty:
        print(f"\n  ⚠  Nenhuma combinação válida encontrada em nenhuma largura de bobina.")
        print(f"     Sugestão: revise os parâmetros ou amplie MAX_COMP_NA_COMBO.")
        print(sep)
        return

    print(f"  Combinações    : {len(df_res)}\n")
    fmt = f"  {{:<5}} {{:<48}} {{:<12}} {{:<12}} {{:<12}} {{}}"
    print(fmt.format('#', 'Combinação', 'Soma (mm)', 'Perda (mm)', 'Perda (%)', 'Status'))
    print(fmt.format('-'*5, '-'*48, '-'*12, '-'*12, '-'*12, '-'*16))
    for i, r in df_res.iterrows():
        print(fmt.format(
            i + 1,
            r['Combinacao'][:47],
            f"{r['Soma_cortes_mm']:.2f}",
            f"{r['Perda_mm']:.3f}",
            f"{r['Perda_pct']:.4f}%",
            r['Status']
        ))
    print(sep)


# ──────────────────────────────────────────────
#  EXPORTAÇÃO XLSX
# ──────────────────────────────────────────────
def exportar_xlsx(df_res: pd.DataFrame, largura: int,
                  ancora: str, esp: float, tipo: str,
                  caminho: str,
                  qtd_bobinas: int = QTD_BOBINAS_PAD,
                  peso_medio_bob: float = PESO_MEDIO_BOB_PAD,
                  limite_cortes: int | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Combinações"
    ws2 = wb.create_sheet("Detalhes")

    AZUL  = "1F4E79"
    CLARO = "BDD7EE"
    VERDE = "E2EFDA"
    CINZA = "F2F2F2"
    BRAN  = "FFFFFF"
    AMAR  = "FFF2CC"
    ROXO  = "EDE7F6"   # destaque para coluna KG

    thin  = Side(style='thin', color='CCCCCC')
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cel(ws, row, col, val, bold=False, bg=BRAN, fg="000000",
            align="left", fmt=None, wrap=False):
        c = ws.cell(row, col, val)
        c.font      = Font(name="Arial", size=9, bold=bold, color=fg)
        c.fill      = PatternFill("solid", start_color=bg, end_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = borda
        if fmt:
            c.number_format = fmt
        return c

    # ── Pré-calcula Peso médio por bobina ──
    # Peso médio = peso informado pelo usuário / quantidade de bobinas
    # (se usuário informou 12000 kg e 4 bobinas → média = 3000 kg/bobina)
    peso_medio_calc = peso_medio_bob / qtd_bobinas

    def calc_kg_combo(detalhes: list[dict]) -> float:
        """KG total da combinação = soma dos KG de cada matriz."""
        total = 0.0
        for det in detalhes:
            total += (peso_medio_calc / largura) * (det['N_cortes'] * det['Desenvolvimento_mm'] * qtd_bobinas)
        return round(total, 2)

    # ── ABA 1: Combinações ──
    ws1.row_dimensions[1].height = 22
    ws1.cell(1, 1, "PLANO DE CORTE — COMBINAÇÕES VÁLIDAS").font = \
        Font(name="Arial", size=13, bold=True, color=AZUL)
    ws1.merge_cells("A1:I1")

    limite_str = str(limite_cortes) if limite_cortes is not None else "Sem limite"
    refilo_min = REFILO_MIN_ATE_3MM if esp <= 3.0 else REFILO_MIN_ACIMA_3MM
    refilo_regra = f"≤ 3.0 mm → {REFILO_MIN_ATE_3MM} mm | > 3.0 mm → {REFILO_MIN_ACIMA_3MM} mm"
    
    params = [
        ("Matriz Âncora",        ancora),
        ("Espessura",            f"{esp} mm"),
        ("Tipo de Material",     tipo),
        ("Largura da Bobina",    f"{largura} mm"),
        ("Padrões Testados",     " → ".join(str(l) for l in LARGURAS_BOBINA) + f"  (usado: {largura} mm)"),
        ("Limite de Cortes",     limite_str),
        ("Refilo Mínimo",        f"{refilo_min} mm  (regra: {refilo_regra})"),
        ("Qtd. de Bobinas",      str(qtd_bobinas)),
        ("Peso Informado",        f"{peso_medio_bob:,.0f} kg  ({peso_medio_bob/1000:.1f} ton)"),
        ("Peso Médio / Bobina",   f"{peso_medio_calc:,.0f} kg  ({peso_medio_calc/1000:.2f} ton)"),
        ("Perda Mínima Aceita",  f"{PERDA_MIN_PCT}%  ({largura * PERDA_MIN_PCT / 100:.2f} mm)"),
        ("Perda Máxima Aceita",  f"{PERDA_MAX_PCT}%  ({largura * PERDA_MAX_PCT / 100:.2f} mm)"),
        ("Total de Combinações", len(df_res)),
    ]
    for r, (k, v) in enumerate(params, start=2):
        cel(ws1, r, 1, k, bold=True, bg=CLARO)
        cel(ws1, r, 2, v, bg=CLARO)
        ws1.merge_cells(f"B{r}:I{r}")

    SR = len(params) + 3
    ws1.row_dimensions[SR].height = 28
    hdrs = ["#", "Combinação  (Âncora em destaque)", "N Âncora",
            "Total Cortes", "Soma Cortes (mm)", "Perda (mm)", "Perda (%)",
            "Qtd. KG", "Status"]
    for col, h in enumerate(hdrs, 1):
        cel(ws1, SR, col, h, bold=True, bg=AZUL, fg=BRAN, align="center")

    for i, row in df_res.iterrows():
        r   = SR + 1 + i
        bg  = VERDE if i % 2 == 0 else CINZA
        kg  = calc_kg_combo(row['Detalhes'])
        
        # Cor do status: verde se válida, laranja se fora da regra
        status_val = row['Status']
        status_bg  = VERDE if status_val == "✓ Válida" else "FFE0B2"  # laranja claro
        
        cel(ws1, r, 1, i + 1,                   bg=bg,        align="center")
        cel(ws1, r, 2, row['Combinacao'],        bg=bg,        wrap=True)
        cel(ws1, r, 3, row['N_ancora'],          bg=AMAR,      align="center")
        cel(ws1, r, 4, row['Total_cortes'],      bg=bg,        align="center")
        cel(ws1, r, 5, row['Soma_cortes_mm'],    bg=bg,        align="right", fmt='#,##0.000')
        cel(ws1, r, 6, row['Perda_mm'],          bg=bg,        align="right", fmt='#,##0.000')
        cel(ws1, r, 7, row['Perda_pct'] / 100,  bg=bg,        align="right", fmt='0.0000%')
        cel(ws1, r, 8, kg,                       bg=ROXO,      align="right", fmt='#,##0.00')
        cel(ws1, r, 9, status_val,               bg=status_bg, align="center")
        ws1.row_dimensions[r].height = 16

    for col, w in zip("ABCDEFGHI", [5, 52, 10, 13, 18, 13, 12, 16, 10]):
        ws1.column_dimensions[col].width = w

    # ── ABA 2: Detalhes (com KG por matriz) ──
    ws2.row_dimensions[1].height = 20
    ws2.cell(1, 1, "DETALHES POR COMBINAÇÃO").font = \
        Font(name="Arial", size=12, bold=True, color=AZUL)
    ws2.merge_cells("A1:G1")

    for col, h in enumerate(
        ["# Combo", "Papel", "Matriz", "Desenvolvimento (mm)", "N° Cortes", "Subtotal (mm)", "Qtd. KG"], 1
    ):
        cel(ws2, 2, col, h, bold=True, bg=AZUL, fg=BRAN, align="center")

    r = 3
    for i, row in df_res.iterrows():
        for j, det in enumerate(row['Detalhes']):
            papel  = "ÂNCORA" if j == 0 else "Complementar"
            bg_det = AMAR if j == 0 else BRAN
            kg_det = round(
                (peso_medio_calc / largura) * (det['N_cortes'] * det['Desenvolvimento_mm'] * qtd_bobinas), 2
            )
            cel(ws2, r, 1, i + 1,                     align="center")
            cel(ws2, r, 2, papel,     bg=bg_det,       align="center", bold=(j == 0))
            cel(ws2, r, 3, det['Matriz'],               bg=bg_det)
            cel(ws2, r, 4, det['Desenvolvimento_mm'],   bg=bg_det, align="right", fmt='#,##0.000')
            cel(ws2, r, 5, det['N_cortes'],             bg=bg_det, align="center")
            cel(ws2, r, 6, det['Subtotal_mm'],          bg=bg_det, align="right", fmt='#,##0.000')
            cel(ws2, r, 7, kg_det,                      bg=ROXO,   align="right", fmt='#,##0.00')
            r += 1

    for col, w in zip("ABCDEFG", [10, 14, 28, 22, 12, 16, 16]):
        ws2.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else ".", exist_ok=True)
    wb.save(caminho)
    print(f"\n  ✓ Resultado exportado: {caminho}")


# ──────────────────────────────────────────────
#  INTERFACE CLI
# ──────────────────────────────────────────────
def menu(df: pd.DataFrame) -> tuple[float, str, str, int | None, int, float]:
    print("\n" + "=" * 62)
    print("          SISTEMA DE PLANO DE CORTE")
    print(f"  Larguras de bobina testadas: {' → '.join(str(l) for l in LARGURAS_BOBINA)} mm")
    print("=" * 62)

    # 1. Espessura
    espessuras = listar_espessuras(df)
    print("\n[1] Espessuras disponíveis:")
    for i, e in enumerate(espessuras, 1):
        print(f"    {i:3d}. {e} mm")
    while True:
        try:
            esp = espessuras[int(input("\n  Número da espessura: ")) - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")

    # 2. Tipo de material
    tipos = listar_tipos(df, esp)
    print(f"\n[2] Tipos de material (esp={esp} mm):")
    for i, t in enumerate(tipos, 1):
        print(f"    {i:3d}. {t}")
    while True:
        try:
            tipo = tipos[int(input("\n  Número do tipo: ")) - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")

    # 3. Matriz âncora
    tab = listar_matrizes(df, esp, tipo)
    print(f"\n[3] Matrizes disponíveis (esp={esp} mm / {tipo}):")
    for i, row in tab.iterrows():
        print(f"    {i+1:3d}. {row['Matriz']:<30s}  dev = {row['Dev_mm']:.1f} mm")
    while True:
        try:
            ancora = tab.iloc[int(input("\n  Número da MATRIZ ÂNCORA: ")) - 1]['Matriz']
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")

    # 4. Limite de cortes (opcional)
    print(f"\n[4] Limite máximo de cortes por combinação (restrição de máquina)")
    print(f"    Deixe em branco e pressione Enter para sem limite.")
    while True:
        raw = input("  Limite de cortes: ").strip()
        if raw == "":
            limite_cortes = None
            break
        try:
            limite_cortes = int(raw)
            if limite_cortes < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número inteiro positivo ou deixe em branco.")

    # 5. Quantidade de bobinas
    print(f"\n[5] Quantidade de bobinas para cálculo de KG")
    while True:
        raw = input(f"  Quantidade de bobinas [{QTD_BOBINAS_PAD}]: ").strip()
        if raw == "":
            qtd_bobinas = QTD_BOBINAS_PAD
            break
        try:
            qtd_bobinas = int(raw)
            if qtd_bobinas < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número inteiro positivo.")

    # 6. Peso médio por bobina
    print(f"\n[6] Peso médio por bobina (kg) para cálculo de KG")
    while True:
        raw = input(f"  Peso médio por bobina (kg) [{PESO_MEDIO_BOB_PAD:,.0f}]: ").strip()
        if raw == "":
            peso_medio_bob = float(PESO_MEDIO_BOB_PAD)
            break
        try:
            peso_medio_bob = float(raw.replace(',', '.').replace('.', '', raw.count('.') - 1))
            if peso_medio_bob <= 0:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo (ex: 12000 ou 15000).")

    return esp, tipo, ancora, limite_cortes, qtd_bobinas, peso_medio_bob


# ──────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────
def main():
    caminho_db = os.path.join(BASE_INPUT, 'db_plano_corte.xlsx')
    print(f"\n  Carregando: {caminho_db}")
    df = carregar_dados(caminho_db)
    print(f"  {len(df)} produtos carregados.")

    esp, tipo, ancora, limite_cortes, qtd_bobinas, peso_medio_bob = menu(df)

    lim_str = f"  Limite cortes  : {limite_cortes}" if limite_cortes else ""
    print(f"\n  Buscando combinações para: {ancora} | {esp} mm | {tipo}{lim_str}")
    df_res, largura = encontrar_combinacoes(
        df, esp, tipo, ancora, MAX_COMP_NA_COMBO, limite_cortes=limite_cortes
    )

    exibir(df_res, largura, ancora, esp, tipo, limite_cortes=limite_cortes)

    if not df_res.empty:
        ancora_safe = ancora.replace('/', '_').replace('"', 'in').replace(',', '-').replace(' ', '_')
        nome = f"plano_{ancora_safe}_esp{str(esp).replace('.', '-')}_{tipo.replace(' ', '_')}_L{largura}.xlsx"
        exportar_xlsx(
            df_res, largura, ancora, esp, tipo,
            os.path.join(BASE_OUTPUT, nome),
            qtd_bobinas=qtd_bobinas,
            peso_medio_bob=peso_medio_bob,
            limite_cortes=limite_cortes,
        )


if __name__ == '__main__':
    main()