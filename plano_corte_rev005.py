"""
================================================================================
PLANO DE CORTE — OTIMIZADOR DE COMBINAÇÕES DE MATRIZES
================================================================================

OBJETIVO:
    Encontrar a melhor forma de combinar diferentes perfis (matrizes) em uma
    bobina de aço, minimizando o desperdício (refilo/perda).

REGRAS DE NEGÓCIO:
    1. Usuário escolhe: espessura + tipo de material + matriz âncora
    2. Matriz âncora é OBRIGATÓRIA (sempre presente com N ≥ 1 corte)
    3. Larguras de bobina fixas: 1200, 1000 ou 1500 mm
    4. Perda aceitável: entre 0,67% e 1,70% da largura
    5. Refilo mínimo por espessura:
       - Espessuras ≤ 3.0 mm: mínimo 10 mm de refilo
       - Espessuras > 3.0 mm: mínimo 14 mm de refilo
    6. Limite de cortes (opcional): máximo de cortes simultâneos

ESTRUTURA DO CÓDIGO:
    BLOCO 1: Configurações e constantes
    BLOCO 2: Funções de carga e limpeza de dados
    BLOCO 3: Funções de consulta ao banco de dados
    BLOCO 4: Motor de busca combinatorial
    BLOCO 5: Cálculo de KG
    BLOCO 6: Validação de resultados
    BLOCO 7: Interface com usuário (CLI)
    BLOCO 8: Exportação para Excel
    BLOCO 9: Função principal (main)
================================================================================
"""

import os
import platform
import pandas as pd
from itertools import combinations, product as iproduct
from datetime import datetime


# ================================================================================
# BLOCO 1: CONFIGURAÇÕES E CONSTANTES
# ================================================================================

# ── Detecta sistema operacional e define caminhos ──
SO = platform.system()
USUARIO = os.getenv('USERNAME') if SO == 'Windows' else os.getenv('USER')

if SO == 'Windows':
    BASE_INPUT  = r'D:\#Mega\Jeferson - Dev\02 - Linguagens\Python\Acotel\Plano_corte_py\files\input'
    BASE_OUTPUT = r'D:\#Mega\Jeferson - Dev\02 - Linguagens\Python\Acotel\Plano_corte_py\files\output'
elif SO == 'Linux':
    BASE_INPUT  = r'/home/stark/Documentos/Dev/Plano_corte_py/files/input'
    BASE_OUTPUT = r'/home/stark/Documentos/Dev/Plano_corte_py/files/output'
else:
    raise Exception(f'Sistema operacional não suportado: {SO}')


# ── Parâmetros de negócio (ÚNICOS VALORES A AJUSTAR) ──
# Larguras de bobina disponíveis (em ordem de preferência)
LARGURAS_BOBINA = [1200, 1000, 1500]

# Janela de perda aceitável (em percentual da largura)
PERDA_MIN_PCT = 0.67
PERDA_MAX_PCT = 1.70

# Refilo mínimo por espessura (em milímetros)
REFILO_MIN_ATE_3MM   = 10   # para espessuras ≤ 3.0 mm
REFILO_MIN_ACIMA_3MM = 14   # para espessuras > 3.0 mm

# Quantidade máxima de matrizes complementares por combinação
MAX_COMP_NA_COMBO = 2

# Valores padrão para cálculo de KG
PESO_MEDIO_BOB_PAD = 12_000  # kg (12 toneladas)
QTD_BOBINAS_PAD = 1


# ================================================================================
# BLOCO 2: FUNÇÕES DE CARGA E LIMPEZA DE DADOS
# ================================================================================

def carregar_dados(caminho: str) -> pd.DataFrame:
    """
    Carrega o arquivo Excel com as matrizes e faz limpeza dos dados.
    
    ENTRADA:
        caminho: path completo do arquivo Excel (db_plano_corte.xlsx)
    
    SAÍDA:
        DataFrame com colunas limpas e validadas:
        - Matriz: nome do perfil (texto limpo)
        - Tipo de material: COMERCIAL, GALVANIZADO, etc
        - Espessura: em mm (número)
        - Desenvolvimento: largura necessária em mm (número)
    
    LIMPEZA REALIZADA:
        1. Remove espaços em branco das strings
        2. Converte espessura e desenvolvimento para números
        3. Remove linhas com dados ausentes ou inválidos
        4. Remove matrizes com desenvolvimento zero ou negativo
    """
    # Lê arquivo Excel
    df = pd.read_excel(caminho)
    
    # Limpa strings (remove espaços invisíveis)
    df['Matriz'] = df['Matriz'].astype(str).str.strip()
    df['Tipo de material'] = df['Tipo de material'].astype(str).str.strip()
    
    # Converte para numérico (valores inválidos viram NaN)
    df['Espessura'] = pd.to_numeric(df['Espessura'], errors='coerce')
    df['Desenvolvimento'] = pd.to_numeric(df['Desenvolvimento'], errors='coerce')
    
    # Remove linhas com campos obrigatórios ausentes
    df = df.dropna(subset=['Espessura', 'Desenvolvimento', 'Matriz', 'Tipo de material'])
    
    # Remove desenvolvimentos inválidos
    df = df[df['Desenvolvimento'] > 0]
    
    # Remove matrizes vazias
    df = df[~df['Matriz'].isin(['nan', ''])]
    
    return df


# ================================================================================
# BLOCO 3: FUNÇÕES DE CONSULTA AO BANCO DE DADOS
# ================================================================================

def listar_espessuras(df: pd.DataFrame) -> list[float]:
    """
    Retorna lista de todas as espessuras disponíveis no banco, ordenadas.
    
    ENTRADA: DataFrame com dados das matrizes
    SAÍDA: Lista de espessuras únicas [0.4, 0.5, ..., 4.75, ...]
    """
    return sorted(df['Espessura'].unique())


def listar_tipos(df: pd.DataFrame, espessura: float) -> list[str]:
    """
    Retorna tipos de material disponíveis para uma espessura específica.
    
    ENTRADA:
        df: DataFrame com dados das matrizes
        espessura: espessura escolhida pelo usuário (ex: 2.0)
    
    SAÍDA:
        Lista de tipos disponíveis ['COMERCIAL', 'GALVANIZADO', ...]
    """
    return sorted(df[df['Espessura'] == espessura]['Tipo de material'].unique())


def listar_matrizes(df: pd.DataFrame, espessura: float, tipo: str) -> pd.DataFrame:
    """
    Retorna todas as matrizes disponíveis para espessura + tipo específicos.
    
    ENTRADA:
        df: DataFrame com dados das matrizes
        espessura: espessura escolhida (ex: 2.0)
        tipo: tipo de material escolhido (ex: 'COMERCIAL')
    
    SAÍDA:
        DataFrame com colunas:
        - Matriz: nome do perfil
        - Dev_mm: desenvolvimento médio em mm
        
    NOTA: Se uma matriz aparece múltiplas vezes no banco (por estar em
          diferentes produtos), calcula a média do desenvolvimento.
    """
    # Filtra por espessura e tipo
    mask = (df['Espessura'] == espessura) & (df['Tipo de material'] == tipo)
    
    # Agrupa por matriz e calcula média do desenvolvimento
    return (
        df[mask]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'Dev_mm'})
        .sort_values('Dev_mm', ascending=False)  # maior primeiro
        .reset_index(drop=True)
    )


def obter_desenvolvimento(df: pd.DataFrame, matriz: str, espessura: float) -> float:
    """
    Obtém o desenvolvimento (largura necessária) de uma matriz específica.
    
    ENTRADA:
        df: DataFrame com dados das matrizes
        matriz: nome da matriz (ex: '50,80-2"')
        espessura: espessura da matriz (ex: 2.0)
    
    SAÍDA:
        Desenvolvimento em mm (ex: 157.0)
    
    ERRO:
        ValueError se a matriz não existir no banco
    """
    # Filtra por matriz e espessura
    mask = (df['Matriz'] == matriz) & (df['Espessura'] == espessura)
    vals = df[mask]['Desenvolvimento'].dropna()
    
    if vals.empty:
        raise ValueError(f"Matriz '{matriz}' com espessura {espessura} mm não encontrada.")
    
    # Retorna média (caso matriz apareça múltiplas vezes)
    return float(vals.mean())


# ================================================================================
# BLOCO 4: MOTOR DE BUSCA COMBINATORIAL
# ================================================================================

def buscar_combinacoes_para_largura(
    dev_ancora: float,
    matriz_ancora: str,
    matrizes_complementares: list[str],
    devs_complementares: list[float],
    largura_bobina: int,
    max_complementares: int,
    espessura: float,
    limite_cortes: int | None = None
) -> list[dict]:
    """
    Motor principal: testa TODAS as combinações possíveis para uma largura.
    
    ALGORITMO:
        Para cada quantidade N de cortes da âncora (1, 2, 3, ...):
            1. Calcula espaço restante = largura - (dev_ancora × N)
            2. Testa âncora sozinha (sem complementares)
            3. Testa âncora + 1 complementar
            4. Testa âncora + 2 complementares
            5. Para cada combinação, valida perda % e refilo
    
    ENTRADA:
        dev_ancora: desenvolvimento da matriz âncora em mm
        matriz_ancora: nome da matriz âncora
        matrizes_complementares: lista de nomes das outras matrizes
        devs_complementares: desenvolvimentos correspondentes
        largura_bobina: 1000, 1200 ou 1500 mm
        max_complementares: quantas complementares permitir (padrão: 2)
        espessura: espessura em mm (para calcular refilo mínimo)
        limite_cortes: soma máxima de cortes permitida (None = sem limite)
    
    SAÍDA:
        Lista de dicionários, cada um representando uma combinação válida:
        {
            'Combinacao': '50,80-2"(x3) + 38,10(x2)',
            'N_ancora': 3,
            'Num_comp': 1,
            'Total_cortes': 5,
            'Detalhes': [...],
            'Soma_cortes_mm': 1191.0,
            'Perda_mm': 9.0,
            'Perda_pct': 0.75,
            'Largura_bobina': 1200,
            'Status': '✓ Válida' ou 'Fora da regra'
        }
    """
    # ── Calcula limites de validação ──
    perda_min_mm = largura_bobina * PERDA_MIN_PCT / 100
    perda_max_mm = largura_bobina * PERDA_MAX_PCT / 100
    
    # Refilo mínimo depende da espessura
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    
    # Máximo de cortes da âncora que cabem na bobina
    max_n_ancora = int(largura_bobina / dev_ancora)
    
    resultados = []
    
    # ── Loop principal: varia quantidade de cortes da âncora ──
    for n_ancora in range(1, max_n_ancora + 1):
        
        # Calcula quanto a âncora ocupa
        soma_ancora = dev_ancora * n_ancora
        espaco_restante = largura_bobina - soma_ancora
        
        # Se âncora já ultrapassou, para
        if espaco_restante < 0:
            break
        
        # ═══════════════════════════════════════════════════════════
        # CASO 1: SÓ A ÂNCORA (sem complementares)
        # ═══════════════════════════════════════════════════════════
        
        perda_mm = espaco_restante
        total_cortes = n_ancora
        
        # Validação em cascata:
        # 1º: perda % deve estar na janela
        passa_pct = (perda_min_mm <= perda_mm <= perda_max_mm)
        
        # 2º: perda mm deve ser >= refilo mínimo
        passa_refilo = (perda_mm >= refilo_min)
        
        # 3º: total de cortes deve respeitar limite
        passa_cortes = (limite_cortes is None or total_cortes <= limite_cortes)
        
        # Se passou na janela de % E no limite de cortes
        if passa_pct and passa_cortes:
            # Define status baseado no refilo
            if passa_refilo:
                status = "✓ Válida"
            else:
                status = "Fora da regra"
            
            # Adiciona aos resultados
            resultados.append({
                'Combinacao': f'{matriz_ancora}(x{n_ancora})',
                'N_ancora': n_ancora,
                'Num_comp': 0,
                'Total_cortes': total_cortes,
                'Detalhes': [{
                    'Matriz': matriz_ancora,
                    'Desenvolvimento_mm': dev_ancora,
                    'N_cortes': n_ancora,
                    'Subtotal_mm': round(soma_ancora, 3)
                }],
                'Soma_cortes_mm': round(soma_ancora, 3),
                'Perda_mm': round(perda_mm, 3),
                'Perda_pct': round(perda_mm / largura_bobina * 100, 4),
                'Largura_bobina': largura_bobina,
                'Status': status
            })
        
        # ═══════════════════════════════════════════════════════════
        # CASO 2: ÂNCORA + COMPLEMENTARES
        # ═══════════════════════════════════════════════════════════
        
        # Se não sobrou espaço para nenhuma complementar, pula
        if espaco_restante < min(devs_complementares, default=largura_bobina + 1):
            continue
        
        # Filtra só as complementares que cabem no espaço restante
        indices_que_cabem = [
            i for i, dev in enumerate(devs_complementares)
            if dev <= espaco_restante
        ]
        
        if not indices_que_cabem:
            continue
        
        # Testa com 1 complementar, depois 2, etc (até max_complementares)
        for qtd_comp in range(1, min(max_complementares, len(indices_que_cabem)) + 1):
            
            # Gera todas as combinações de 'qtd_comp' matrizes
            # Ex: se qtd_comp=2 e temos [A,B,C], gera: (A,B), (A,C), (B,C)
            for indices_escolhidos in combinations(indices_que_cabem, qtd_comp):
                
                # Pega desenvolvimentos e nomes das escolhidas
                devs = [devs_complementares[i] for i in indices_escolhidos]
                nomes = [matrizes_complementares[i] for i in indices_escolhidos]
                
                # Para cada matriz, calcula quantos cortes cabem
                max_cortes_cada = [max(1, int(espaco_restante / d)) for d in devs]
                
                # Gera todas as combinações de quantidades de cortes
                # Ex: se max_cortes_cada = [3, 2], gera:
                #     (1,1), (1,2), (2,1), (2,2), (3,1), (3,2)
                for qtds_cortes in iproduct(*[range(1, mx + 1) for mx in max_cortes_cada]):
                    
                    # Calcula soma das complementares
                    soma_comp = sum(d * n for d, n in zip(devs, qtds_cortes))
                    soma_total = soma_ancora + soma_comp
                    total_cortes = n_ancora + sum(qtds_cortes)
                    
                    # Se ultrapassou a largura, pula
                    if soma_total > largura_bobina:
                        continue
                    
                    # Se ultrapassou limite de cortes, pula
                    if limite_cortes is not None and total_cortes > limite_cortes:
                        continue
                    
                    # Calcula perda
                    perda_mm = largura_bobina - soma_total
                    
                    # Validação em cascata
                    passa_pct = (perda_min_mm <= perda_mm <= perda_max_mm)
                    passa_refilo = (perda_mm >= refilo_min)
                    
                    # Se passou na janela de %
                    if passa_pct:
                        # Define status baseado no refilo
                        if passa_refilo:
                            status = "✓ Válida"
                        else:
                            status = "Fora da regra"
                        
                        # Monta lista de detalhes (âncora + cada complementar)
                        detalhes = [{
                            'Matriz': matriz_ancora,
                            'Desenvolvimento_mm': dev_ancora,
                            'N_cortes': n_ancora,
                            'Subtotal_mm': round(soma_ancora, 3)
                        }]
                        
                        for nome, dev, n in zip(nomes, devs, qtds_cortes):
                            detalhes.append({
                                'Matriz': nome,
                                'Desenvolvimento_mm': dev,
                                'N_cortes': n,
                                'Subtotal_mm': round(dev * n, 3)
                            })
                        
                        # Monta string da combinação
                        comp_str = ' + '.join(f'{nome}(x{n})' for nome, n in zip(nomes, qtds_cortes))
                        
                        # Adiciona aos resultados
                        resultados.append({
                            'Combinacao': f'{matriz_ancora}(x{n_ancora}) + {comp_str}',
                            'N_ancora': n_ancora,
                            'Num_comp': qtd_comp,
                            'Total_cortes': total_cortes,
                            'Detalhes': detalhes,
                            'Soma_cortes_mm': round(soma_total, 3),
                            'Perda_mm': round(perda_mm, 3),
                            'Perda_pct': round(perda_mm / largura_bobina * 100, 4),
                            'Largura_bobina': largura_bobina,
                            'Status': status
                        })
    
    return resultados


def encontrar_combinacoes(
    df: pd.DataFrame,
    espessura: float,
    tipo_material: str,
    matriz_ancora: str,
    limite_cortes: int | None = None
) -> tuple[pd.DataFrame, int]:
    """
    Orquestrador principal: tenta larguras em sequência até encontrar resultado.
    
    ESTRATÉGIA:
        1. Tenta largura 1200 mm (mais comum)
        2. Se não houver resultados, tenta 1000 mm
        3. Se ainda não houver, tenta 1500 mm
        4. Para na primeira que retornar combinações válidas
    
    ENTRADA:
        df: DataFrame com todas as matrizes
        espessura: espessura escolhida pelo usuário
        tipo_material: tipo escolhido pelo usuário
        matriz_ancora: matriz âncora escolhida pelo usuário
        limite_cortes: limite opcional de cortes totais
    
    SAÍDA:
        (DataFrame com resultados, largura_usada)
        
        Se nenhuma largura retornar resultados: (DataFrame vazio, 0)
    """
    # ── Pega desenvolvimento da âncora ──
    dev_ancora = obter_desenvolvimento(df, matriz_ancora, espessura)
    
    # ── Busca matrizes complementares (mesma espessura + tipo, exceto âncora) ──
    mask_comp = (
        (df['Espessura'] == espessura) &
        (df['Tipo de material'] == tipo_material) &
        (df['Matriz'] != matriz_ancora)
    )
    
    candidatas = (
        df[mask_comp]
        .groupby('Matriz')['Desenvolvimento']
        .mean()
        .reset_index()
        .rename(columns={'Desenvolvimento': 'dev'})
        .sort_values('dev', ascending=False)  # maior primeiro (otimização)
    )
    
    matrizes_comp = candidatas['Matriz'].tolist()
    devs_comp = candidatas['dev'].tolist()
    
    # ── Tenta cada largura em ordem ──
    for largura in LARGURAS_BOBINA:
        print(f"  → Tentando largura {largura} mm ...", end=' ')
        
        # Verifica se âncora cabe ao menos uma vez
        if dev_ancora > largura:
            print(f"âncora ({dev_ancora:.1f}mm) não cabe. Pulando.")
            continue
        
        # Chama motor de busca
        resultados = buscar_combinacoes_para_largura(
            dev_ancora=dev_ancora,
            matriz_ancora=matriz_ancora,
            matrizes_complementares=matrizes_comp,
            devs_complementares=devs_comp,
            largura_bobina=largura,
            max_complementares=MAX_COMP_NA_COMBO,
            espessura=espessura,
            limite_cortes=limite_cortes
        )
        
        # Se encontrou resultados, para aqui
        if resultados:
            print(f"{len(resultados)} combinações encontradas. ✓")
            
            # Converte para DataFrame e ordena
            df_res = (
                pd.DataFrame(resultados)
                .sort_values(['Perda_pct', 'N_ancora', 'Num_comp'])
                .reset_index(drop=True)
            )
            
            return df_res, largura
        else:
            print("nenhuma combinação válida.")
    
    # Se chegou aqui, nenhuma largura teve resultado
    return pd.DataFrame(), 0


# ================================================================================
# BLOCO 5: CÁLCULO DE KG
# ================================================================================

def calcular_peso_medio_bobina(peso_informado: float, qtd_bobinas: int) -> float:
    """
    Calcula o peso médio por bobina.
    
    ENTRADA:
        peso_informado: peso TOTAL do lote em kg (ex: 48.000 kg)
        qtd_bobinas: quantidade de bobinas no lote (ex: 4)
    
    SAÍDA:
        Peso médio por bobina em kg (ex: 12.000 kg)
    
    EXEMPLO:
        48.000 kg ÷ 4 bobinas = 12.000 kg/bobina
    """
    return peso_informado / qtd_bobinas


def calcular_kg_matriz(
    peso_medio_bobina: float,
    largura_bobina: int,
    n_cortes: int,
    desenvolvimento: float,
    qtd_bobinas: int
) -> float:
    """
    Calcula KG de aço para uma matriz específica dentro da combinação.
    
    FÓRMULA:
        KG = (Peso_médio / Largura) × (N_cortes × Desenvolvimento × Qtd_bobinas)
    
    ENTRADA:
        peso_medio_bobina: peso médio calculado (kg/bobina)
        largura_bobina: 1000, 1200 ou 1500 mm
        n_cortes: quantos cortes desta matriz
        desenvolvimento: largura necessária em mm
        qtd_bobinas: quantidade de bobinas no lote
    
    SAÍDA:
        Quilos de aço para esta matriz
    
    EXEMPLO:
        peso_medio = 12.000 kg
        largura = 1.200 mm
        n_cortes = 3
        desenvolvimento = 157 mm
        qtd_bobinas = 4
        
        KG = (12.000 / 1.200) × (3 × 157 × 4)
        KG = 10 × 1.884
        KG = 18.840 kg
    """
    return (peso_medio_bobina / largura_bobina) * (n_cortes * desenvolvimento * qtd_bobinas)


def calcular_kg_combinacao(
    detalhes_combinacao: list[dict],
    peso_medio_bobina: float,
    largura_bobina: int,
    qtd_bobinas: int
) -> float:
    """
    Calcula KG total de uma combinação (soma dos KGs de cada matriz).
    
    ENTRADA:
        detalhes_combinacao: lista de dicts com info de cada matriz
        peso_medio_bobina: peso médio por bobina em kg
        largura_bobina: largura usada em mm
        qtd_bobinas: quantidade de bobinas
    
    SAÍDA:
        KG total da combinação
    """
    total_kg = 0.0
    
    for matriz_info in detalhes_combinacao:
        kg_matriz = calcular_kg_matriz(
            peso_medio_bobina=peso_medio_bobina,
            largura_bobina=largura_bobina,
            n_cortes=matriz_info['N_cortes'],
            desenvolvimento=matriz_info['Desenvolvimento_mm'],
            qtd_bobinas=qtd_bobinas
        )
        total_kg += kg_matriz
    
    return round(total_kg, 2)


# ================================================================================
# BLOCO 6: VALIDAÇÃO DE RESULTADOS
# ================================================================================

def validar_resultado(df_res: pd.DataFrame, espessura: float) -> dict:
    """
    Valida o resultado e retorna estatísticas.
    
    ENTRADA:
        df_res: DataFrame com combinações encontradas
        espessura: espessura usada (para determinar refilo)
    
    SAÍDA:
        Dicionário com estatísticas:
        {
            'total': 104,
            'validas': 102,
            'fora_regra': 2,
            'refilo_min': 10
        }
    """
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    
    validas = len(df_res[df_res['Status'] == '✓ Válida'])
    fora_regra = len(df_res[df_res['Status'] == 'Fora da regra'])
    
    return {
        'total': len(df_res),
        'validas': validas,
        'fora_regra': fora_regra,
        'refilo_min': refilo_min
    }


# ================================================================================
# BLOCO 7: INTERFACE COM USUÁRIO (CLI)
# ================================================================================

def exibir_terminal(
    df_res: pd.DataFrame,
    largura: int,
    ancora: str,
    espessura: float,
    tipo: str,
    limite_cortes: int | None = None
) -> None:
    """
    Exibe resultados formatados no terminal.
    
    ENTRADA:
        df_res: DataFrame com combinações
        largura: largura da bobina usada
        ancora: nome da matriz âncora
        espessura: espessura em mm
        tipo: tipo de material
        limite_cortes: limite opcional de cortes
    """
    sep = "=" * 90
    print(f"\n{sep}")
    print("  PLANO DE CORTE — COMBINAÇÕES VÁLIDAS")
    print(sep)
    print(f"  Âncora         : {ancora}")
    print(f"  Espessura      : {espessura} mm")
    print(f"  Tipo material  : {tipo}")
    print(f"  Largura bobina : {largura} mm")
    
    # Mostra janela de perda %
    perda_min_mm = largura * PERDA_MIN_PCT / 100
    perda_max_mm = largura * PERDA_MAX_PCT / 100
    print(f"  Janela de perda: {PERDA_MIN_PCT}% – {PERDA_MAX_PCT}%  "
          f"|  {perda_min_mm:.2f} mm – {perda_max_mm:.2f} mm")
    
    # Mostra refilo mínimo
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    simbolo = '≤' if espessura <= 3.0 else '>'
    print(f"  Refilo mínimo  : {refilo_min} mm  (regra para esp {simbolo} 3.0 mm)")
    
    # Mostra limite de cortes se informado
    if limite_cortes is not None:
        print(f"  Limite cortes  : {limite_cortes} cortes (soma total)")
    
    # Se não encontrou nada
    if df_res.empty:
        print(f"\n  ⚠  Nenhuma combinação válida encontrada.")
        print(f"     Sugestão: amplie os parâmetros ou use outra âncora.")
        print(sep)
        return
    
    # Estatísticas
    stats = validar_resultado(df_res, espessura)
    print(f"  Combinações    : {stats['total']} ({stats['validas']} válidas + {stats['fora_regra']} fora da regra)\n")
    
    # Tabela
    fmt = "  {:<5} {:<48} {:<12} {:<12} {:<12} {}"
    print(fmt.format('#', 'Combinação', 'Soma (mm)', 'Perda (mm)', 'Perda (%)', 'Status'))
    print(fmt.format('-'*5, '-'*48, '-'*12, '-'*12, '-'*12, '-'*16))
    
    for i, r in df_res.iterrows():
        print(fmt.format(
            i + 1,
            r['Combinacao'][:47],
            f"{r['Soma_cortes_mm']:.2f}",
            f"{r['Perda_mm']:.3f}",
            f"{r['Perda_pct']:.4f}%",
            r['Status']
        ))
    
    print(sep)


def menu_usuario(df: pd.DataFrame) -> tuple[float, str, str, int | None, int, float]:
    """
    Interface CLI: coleta todas as informações do usuário em 6 passos.
    
    PASSOS:
        [1] Escolhe espessura
        [2] Escolhe tipo de material
        [3] Escolhe matriz âncora
        [4] Informa limite de cortes (opcional)
        [5] Informa quantidade de bobinas
        [6] Informa peso total das bobinas
    
    SAÍDA:
        (espessura, tipo, ancora, limite_cortes, qtd_bobinas, peso_total)
    """
    print("\n" + "=" * 70)
    print("          SISTEMA DE PLANO DE CORTE")
    print(f"  Larguras testadas: {' → '.join(str(l) for l in LARGURAS_BOBINA)} mm")
    print("=" * 70)
    
    # ── [1] Espessura ──
    espessuras = listar_espessuras(df)
    print("\n[1] Espessuras disponíveis:")
    for i, e in enumerate(espessuras, 1):
        print(f"    {i:3d}. {e} mm")
    
    while True:
        try:
            escolha = int(input("\n  Número da espessura: "))
            espessura = espessuras[escolha - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")
    
    # ── [2] Tipo de material ──
    tipos = listar_tipos(df, espessura)
    print(f"\n[2] Tipos de material (esp={espessura} mm):")
    for i, t in enumerate(tipos, 1):
        print(f"    {i:3d}. {t}")
    
    while True:
        try:
            escolha = int(input("\n  Número do tipo: "))
            tipo = tipos[escolha - 1]
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")
    
    # ── [3] Matriz âncora ──
    matrizes = listar_matrizes(df, espessura, tipo)
    print(f"\n[3] Matrizes disponíveis (esp={espessura} mm / {tipo}):")
    for i, row in matrizes.iterrows():
        print(f"    {i+1:3d}. {row['Matriz']:<30s}  dev = {row['Dev_mm']:.1f} mm")
    
    while True:
        try:
            escolha = int(input("\n  Número da MATRIZ ÂNCORA: "))
            ancora = matrizes.iloc[escolha - 1]['Matriz']
            break
        except (ValueError, IndexError):
            print("  ⚠ Inválido. Tente novamente.")
    
    # ── [4] Limite de cortes (opcional) ──
    print(f"\n[4] Limite de cortes por combinação (restrição de máquina)")
    print(f"    Deixe em branco para sem limite.")
    
    while True:
        entrada = input("  Limite de cortes: ").strip()
        if entrada == "":
            limite_cortes = None
            break
        try:
            limite_cortes = int(entrada)
            if limite_cortes < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo ou deixe em branco.")
    
    # ── [5] Quantidade de bobinas ──
    print(f"\n[5] Quantidade de bobinas")
    while True:
        entrada = input(f"  Quantidade [{QTD_BOBINAS_PAD}]: ").strip()
        if entrada == "":
            qtd_bobinas = QTD_BOBINAS_PAD
            break
        try:
            qtd_bobinas = int(entrada)
            if qtd_bobinas < 1:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo.")
    
    # ── [6] Peso total das bobinas ──
    print(f"\n[6] Peso TOTAL do lote de bobinas (kg)")
    while True:
        entrada = input(f"  Peso total (kg) [{PESO_MEDIO_BOB_PAD:,.0f}]: ").strip()
        if entrada == "":
            peso_total = float(PESO_MEDIO_BOB_PAD)
            break
        try:
            # Remove separadores de milhar se houver
            peso_total = float(entrada.replace(',', '.').replace('.', '', entrada.count('.') - 1))
            if peso_total <= 0:
                raise ValueError
            break
        except ValueError:
            print("  ⚠ Digite um número positivo (ex: 48000).")
    
    return espessura, tipo, ancora, limite_cortes, qtd_bobinas, peso_total


# ================================================================================
# BLOCO 8: EXPORTAÇÃO PARA EXCEL
# ================================================================================

def exportar_excel(
    df_res: pd.DataFrame,
    largura: int,
    ancora: str,
    espessura: float,
    tipo: str,
    caminho: str,
    qtd_bobinas: int,
    peso_total: float,
    limite_cortes: int | None = None
) -> None:
    """
    Exporta resultados para arquivo Excel com 2 abas.
    
    ABA 1 - Combinações:
        Resumo de cada combinação com:
        - Cabeçalho com parâmetros usados
        - Tabela com todas as combinações
        - Colunas: #, Combinação, N Âncora, Total Cortes, Soma, Perda, KG, Status
    
    ABA 2 - Detalhes:
        Detalhamento matriz por matriz:
        - Cada linha = uma matriz dentro de uma combinação
        - Identifica âncora vs complementar
        - Mostra KG individual de cada matriz
    
    CORES:
        - Cabeçalho: azul escuro
        - Âncora: amarelo
        - KG: roxo claro
        - Status válida: verde
        - Status fora da regra: laranja
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # ── Cria workbook ──
    wb = Workbook()
    ws_combos = wb.active
    ws_combos.title = "Combinações"
    ws_detalhes = wb.create_sheet("Detalhes")
    
    # ── Paleta de cores ──
    COR_AZUL_ESC = "1F4E79"
    COR_AZUL_CLA = "BDD7EE"
    COR_VERDE    = "E2EFDA"
    COR_CINZA    = "F2F2F2"
    COR_BRANCO   = "FFFFFF"
    COR_AMARELO  = "FFF2CC"
    COR_ROXO     = "EDE7F6"
    COR_LARANJA  = "FFE0B2"
    
    # ── Estilo de borda ──
    borda_fina = Side(style='thin', color='CCCCCC')
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    
    # ── Função auxiliar para criar célula estilizada ──
    def criar_celula(ws, linha, coluna, valor, negrito=False, cor_fundo=COR_BRANCO,
                     cor_texto="000000", alinhamento="left", formato=None, quebra=False):
        """Helper para criar célula com estilo completo"""
        c = ws.cell(linha, coluna, valor)
        c.font = Font(name="Arial", size=9, bold=negrito, color=cor_texto)
        c.fill = PatternFill("solid", start_color=cor_fundo, end_color=cor_fundo)
        c.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=quebra)
        c.border = borda
        if formato:
            c.number_format = formato
        return c
    
    # ── Calcula peso médio ──
    peso_medio = calcular_peso_medio_bobina(peso_total, qtd_bobinas)
    
    # ══════════════════════════════════════════════════════════════
    # ABA 1: COMBINAÇÕES
    # ══════════════════════════════════════════════════════════════
    
    # ── Título ──
    ws_combos.row_dimensions[1].height = 22
    ws_combos.cell(1, 1, "PLANO DE CORTE — COMBINAÇÕES VÁLIDAS").font = \
        Font(name="Arial", size=13, bold=True, color=COR_AZUL_ESC)
    ws_combos.merge_cells("A1:I1")
    
    # ── Cabeçalho de parâmetros ──
    refilo_min = REFILO_MIN_ATE_3MM if espessura <= 3.0 else REFILO_MIN_ACIMA_3MM
    regra_refilo = f"≤ 3.0 mm → {REFILO_MIN_ATE_3MM} mm | > 3.0 mm → {REFILO_MIN_ACIMA_3MM} mm"
    limite_str = str(limite_cortes) if limite_cortes is not None else "Sem limite"
    
    parametros = [
        ("Matriz Âncora", ancora),
        ("Espessura", f"{espessura} mm"),
        ("Tipo de Material", tipo),
        ("Largura da Bobina", f"{largura} mm"),
        ("Padrões Testados", " → ".join(str(l) for l in LARGURAS_BOBINA) + f"  (usado: {largura} mm)"),
        ("Limite de Cortes", limite_str),
        ("Refilo Mínimo", f"{refilo_min} mm  (regra: {regra_refilo})"),
        ("Qtd. de Bobinas", str(qtd_bobinas)),
        ("Peso Total Lote", f"{peso_total:,.0f} kg  ({peso_total/1000:.1f} ton)"),
        ("Peso Médio/Bobina", f"{peso_medio:,.0f} kg  ({peso_medio/1000:.2f} ton)"),
        ("Perda Mínima (%)", f"{PERDA_MIN_PCT}%  ({largura * PERDA_MIN_PCT / 100:.2f} mm)"),
        ("Perda Máxima (%)", f"{PERDA_MAX_PCT}%  ({largura * PERDA_MAX_PCT / 100:.2f} mm)"),
        ("Total Combinações", len(df_res)),
    ]
    
    for r, (chave, valor) in enumerate(parametros, start=2):
        criar_celula(ws_combos, r, 1, chave, negrito=True, cor_fundo=COR_AZUL_CLA)
        criar_celula(ws_combos, r, 2, valor, cor_fundo=COR_AZUL_CLA)
        ws_combos.merge_cells(f"B{r}:I{r}")
    
    # ── Cabeçalho da tabela ──
    linha_cabecalho = len(parametros) + 3
    ws_combos.row_dimensions[linha_cabecalho].height = 28
    
    colunas = ["#", "Combinação  (Âncora em destaque)", "N Âncora", "Total Cortes",
               "Soma Cortes (mm)", "Perda (mm)", "Perda (%)", "Qtd. KG", "Status"]
    
    for col, titulo in enumerate(colunas, 1):
        criar_celula(ws_combos, linha_cabecalho, col, titulo,
                    negrito=True, cor_fundo=COR_AZUL_ESC, cor_texto=COR_BRANCO, alinhamento="center")
    
    # ── Dados das combinações ──
    for i, row in df_res.iterrows():
        linha = linha_cabecalho + 1 + i
        cor_zebra = COR_VERDE if i % 2 == 0 else COR_CINZA
        
        # Calcula KG total
        kg = calcular_kg_combinacao(row['Detalhes'], peso_medio, largura, qtd_bobinas)
        
        # Cor do status
        cor_status = COR_VERDE if row['Status'] == "✓ Válida" else COR_LARANJA
        
        # Preenche células
        criar_celula(ws_combos, linha, 1, i + 1, cor_fundo=cor_zebra, alinhamento="center")
        criar_celula(ws_combos, linha, 2, row['Combinacao'], cor_fundo=cor_zebra, quebra=True)
        criar_celula(ws_combos, linha, 3, row['N_ancora'], cor_fundo=COR_AMARELO, alinhamento="center")
        criar_celula(ws_combos, linha, 4, row['Total_cortes'], cor_fundo=cor_zebra, alinhamento="center")
        criar_celula(ws_combos, linha, 5, row['Soma_cortes_mm'], cor_fundo=cor_zebra, alinhamento="right", formato='#,##0.000')
        criar_celula(ws_combos, linha, 6, row['Perda_mm'], cor_fundo=cor_zebra, alinhamento="right", formato='#,##0.000')
        criar_celula(ws_combos, linha, 7, row['Perda_pct'] / 100, cor_fundo=cor_zebra, alinhamento="right", formato='0.0000%')
        criar_celula(ws_combos, linha, 8, kg, cor_fundo=COR_ROXO, alinhamento="right", formato='#,##0.00')
        criar_celula(ws_combos, linha, 9, row['Status'], cor_fundo=cor_status, alinhamento="center")
        
        ws_combos.row_dimensions[linha].height = 16
    
    # ── Ajusta larguras das colunas ──
    for col, largura_col in zip("ABCDEFGHI", [5, 52, 10, 13, 18, 13, 12, 16, 10]):
        ws_combos.column_dimensions[col].width = largura_col
    
    # ══════════════════════════════════════════════════════════════
    # ABA 2: DETALHES
    # ══════════════════════════════════════════════════════════════
    
    # ── Título ──
    ws_detalhes.row_dimensions[1].height = 20
    ws_detalhes.cell(1, 1, "DETALHES POR COMBINAÇÃO").font = \
        Font(name="Arial", size=12, bold=True, color=COR_AZUL_ESC)
    ws_detalhes.merge_cells("A1:G1")
    
    # ── Cabeçalho ──
    colunas_det = ["# Combo", "Papel", "Matriz", "Desenvolvimento (mm)",
                   "N° Cortes", "Subtotal (mm)", "Qtd. KG"]
    
    for col, titulo in enumerate(colunas_det, 1):
        criar_celula(ws_detalhes, 2, col, titulo,
                    negrito=True, cor_fundo=COR_AZUL_ESC, cor_texto=COR_BRANCO, alinhamento="center")
    
    # ── Dados detalhados ──
    linha = 3
    for i, row in df_res.iterrows():
        for j, detalhe in enumerate(row['Detalhes']):
            # Primeira matriz é sempre a âncora
            papel = "ÂNCORA" if j == 0 else "Complementar"
            cor_papel = COR_AMARELO if j == 0 else COR_BRANCO
            negrito_papel = (j == 0)
            
            # Calcula KG desta matriz
            kg_matriz = calcular_kg_matriz(
                peso_medio, largura,
                detalhe['N_cortes'],
                detalhe['Desenvolvimento_mm'],
                qtd_bobinas
            )
            
            # Preenche linha
            criar_celula(ws_detalhes, linha, 1, i + 1, alinhamento="center")
            criar_celula(ws_detalhes, linha, 2, papel, cor_fundo=cor_papel, alinhamento="center", negrito=negrito_papel)
            criar_celula(ws_detalhes, linha, 3, detalhe['Matriz'], cor_fundo=cor_papel)
            criar_celula(ws_detalhes, linha, 4, detalhe['Desenvolvimento_mm'], cor_fundo=cor_papel, alinhamento="right", formato='#,##0.000')
            criar_celula(ws_detalhes, linha, 5, detalhe['N_cortes'], cor_fundo=cor_papel, alinhamento="center")
            criar_celula(ws_detalhes, linha, 6, detalhe['Subtotal_mm'], cor_fundo=cor_papel, alinhamento="right", formato='#,##0.000')
            criar_celula(ws_detalhes, linha, 7, kg_matriz, cor_fundo=COR_ROXO, alinhamento="right", formato='#,##0.00')
            
            linha += 1
    
    # ── Ajusta larguras ──
    for col, largura_col in zip("ABCDEFG", [10, 14, 28, 22, 12, 16, 16]):
        ws_detalhes.column_dimensions[col].width = largura_col
    
    # ── Salva arquivo ──
    os.makedirs(os.path.dirname(caminho) if os.path.dirname(caminho) else ".", exist_ok=True)
    wb.save(caminho)
    print(f"\n  ✓ Resultado exportado: {caminho}")


# ================================================================================
# BLOCO 9: FUNÇÃO PRINCIPAL (MAIN)
# ================================================================================

def main():
    """
    Função principal que coordena toda a execução.
    
    FLUXO:
        1. Carrega banco de dados
        2. Coleta informações do usuário via menu
        3. Busca combinações válidas
        4. Exibe resultados no terminal
        5. Exporta para Excel
    """
    # ── Carrega banco de dados ──
    caminho_db = os.path.join(BASE_INPUT, 'db_plano_corte.xlsx')
    print(f"\n  Carregando: {caminho_db}")
    
    df = carregar_dados(caminho_db)
    print(f"  ✓ {len(df)} produtos carregados.")
    
    # ── Interface com usuário ──
    espessura, tipo, ancora, limite_cortes, qtd_bobinas, peso_total = menu_usuario(df)
    
    # ── Busca combinações ──
    print(f"\n  Buscando combinações para:")
    print(f"    Âncora: {ancora} | Espessura: {espessura} mm | Tipo: {tipo}")
    if limite_cortes:
        print(f"    Limite de cortes: {limite_cortes}")
    
    df_resultados, largura_usada = encontrar_combinacoes(
        df=df,
        espessura=espessura,
        tipo_material=tipo,
        matriz_ancora=ancora,
        limite_cortes=limite_cortes
    )
    
    # ── Exibe no terminal ──
    exibir_terminal(
        df_res=df_resultados,
        largura=largura_usada,
        ancora=ancora,
        espessura=espessura,
        tipo=tipo,
        limite_cortes=limite_cortes
    )
    
    # ── Exporta para Excel (se houver resultados) ──
    if not df_resultados.empty:
        # Monta nome do arquivo
        ancora_safe = ancora.replace('/', '_').replace('"', 'in').replace(',', '-').replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        esp_str = str(espessura).replace('.', '-')
        tipo_str = tipo.replace(' ', '_')
        
        nome_arquivo = f"plano_{ancora_safe}_esp{esp_str}_{tipo_str}_L{largura_usada}_{timestamp}.xlsx"
        caminho_completo = os.path.join(BASE_OUTPUT, nome_arquivo)
        
        # Exporta
        exportar_excel(
            df_res=df_resultados,
            largura=largura_usada,
            ancora=ancora,
            espessura=espessura,
            tipo=tipo,
            caminho=caminho_completo,
            qtd_bobinas=qtd_bobinas,
            peso_total=peso_total,
            limite_cortes=limite_cortes
        )


# ════════════════════════════════════════════════════════════════════════════════
# EXECUÇÃO
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print(f"Usuário: {USUARIO}")
    main()